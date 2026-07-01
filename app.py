from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify, send_from_directory, make_response
from werkzeug.middleware.proxy_fix import ProxyFix
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, case
import io, hashlib, calendar, os
from datetime import date, datetime, timedelta
from functools import wraps
from waitress import serve
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
app.secret_key = 'attendance_secret_key_2024'

# ── Database Configuration (Auto-detects Render Cloud environment) ────────────
database_url = os.environ.get('DATABASE_URL')
if database_url:
    # Render provides connection strings prefixed with postgres://
    # SQLAlchemy 1.4+ strictly requires postgresql://
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    # Fallback to Microsoft SQL Server for local offline development
    app.config['SQLALCHEMY_DATABASE_URI'] = (
        'mssql+pyodbc://@localhost/attendance_db?driver=ODBC+Driver+18+for+SQL+Server&TrustServerCertificate=yes'
    )
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_recycle': 280,
    'pool_pre_ping': True,
}

db = SQLAlchemy(app)

# ── Models ─────────────────────────────────────────────────────────────────────

class User(db.Model):
    __tablename__ = 'users'
    id          = db.Column(db.Integer, primary_key=True)
    username    = db.Column(db.String(80), unique=True, nullable=False)
    password    = db.Column(db.String(256), nullable=False)
    role        = db.Column(db.String(20), default='employee')
    full_name   = db.Column(db.String(120), nullable=False)
    department  = db.Column(db.String(80), default='General')
    email       = db.Column(db.String(120))
    employee_id = db.Column(db.String(20), db.ForeignKey('employees.employee_id'), nullable=True)

class Employee(db.Model):
    __tablename__ = 'employees'
    id          = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.String(20), unique=True, nullable=False)
    full_name   = db.Column(db.String(120), nullable=False)
    department  = db.Column(db.String(80), nullable=False)
    position    = db.Column(db.String(80))
    email       = db.Column(db.String(120))
    phone       = db.Column(db.String(20))
    join_date   = db.Column(db.Date)
    active      = db.Column(db.Boolean, default=True)
    attendance  = db.relationship('Attendance', backref='employee', lazy='dynamic')

class Attendance(db.Model):
    __tablename__ = 'attendance'
    __table_args__ = (
        db.UniqueConstraint('employee_id', 'date', name='uq_emp_date'),
    )
    id          = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.String(20), db.ForeignKey('employees.employee_id'), nullable=False)
    date        = db.Column(db.Date, nullable=False)
    status      = db.Column(db.String(20), nullable=False)
    check_in    = db.Column(db.String(10))
    check_out   = db.Column(db.String(10))
    notes       = db.Column(db.String(255))
    marked_by   = db.Column(db.String(80))
    location    = db.Column(db.String(100))

class Location(db.Model):
    __tablename__ = 'locations'
    id             = db.Column(db.Integer, primary_key=True)
    name           = db.Column(db.String(100), unique=True, nullable=False)
    latitude       = db.Column(db.Float, nullable=True)   # GPS latitude
    longitude      = db.Column(db.Float, nullable=True)   # GPS longitude
    radius_meters  = db.Column(db.Integer, default=200)   # Geo-fence radius

# ── DB Init & Seed ─────────────────────────────────────────────────────────────

def init_db():
    db.create_all()
    # ── Add columns for users / attendance if using MSSQL (safe legacy migration) ──
    if db.engine.dialect.name == 'mssql':
        with db.engine.connect() as conn:
            conn.execute(db.text("""
                IF NOT EXISTS (
                    SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_NAME = 'users' AND COLUMN_NAME = 'employee_id'
                )
                BEGIN
                    ALTER TABLE users
                    ADD employee_id VARCHAR(20) NULL
                        REFERENCES employees(employee_id)
                END
            """))
            conn.execute(db.text("""
                IF NOT EXISTS (
                    SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_NAME = 'attendance' AND COLUMN_NAME = 'location'
                )
                BEGIN
                    ALTER TABLE attendance
                    ADD location VARCHAR(100) NULL
                END
            """))
            # ── Geo-fence columns for locations table ──────────────────────
            conn.execute(db.text("""
                IF NOT EXISTS (
                    SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_NAME = 'locations' AND COLUMN_NAME = 'latitude'
                )
                BEGIN
                    ALTER TABLE locations ADD latitude FLOAT NULL
                END
            """))
            conn.execute(db.text("""
                IF NOT EXISTS (
                    SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_NAME = 'locations' AND COLUMN_NAME = 'longitude'
                )
                BEGIN
                    ALTER TABLE locations ADD longitude FLOAT NULL
                END
            """))
            conn.execute(db.text("""
                IF NOT EXISTS (
                    SELECT 1 FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE TABLE_NAME = 'locations' AND COLUMN_NAME = 'radius_meters'
                )
                BEGIN
                    ALTER TABLE locations ADD radius_meters INT DEFAULT 200
                END
            """))
            conn.commit()
    # ──────────────────────────────────────────────────────────────────────────
    # Seed admin
    if not User.query.filter_by(username='admin').first():
        pw = hashlib.sha256('admin123'.encode()).hexdigest()
        db.session.add(User(username='admin', password=pw, role='admin',
                            full_name='System Admin', department='IT'))
        db.session.commit()
    # Seed actual employees
    samples = [
        ('EMP012', 'Arun Venkachatalam', 'Management', 'Founder', 'arun@jdtech.com', '', date(2026, 1, 1)),
        ('EMP013', 'Vinoth Kumar', 'SCADA', 'SCADA Engineer', 'vinoth@jdtech.com', '', date(2026, 1, 1)),
        ('EMP014', 'Ajith', 'SCADA', 'SCADA Engineer', 'ajith@jdtech.com', '', date(2026, 1, 1)),
        ('EMP015', 'Mano Ranjith', 'SCADA', 'SCADA Engineer', 'manoranjith@jdtech.com', '', date(2026, 1, 1)),
        ('EMP016', 'Praveen', 'SCADA', 'SCADA Engineer', 'praveen@jdtech.com', '', date(2026, 1, 1)),
        ('EMP017', 'Arivu', 'SCADA', 'SCADA Engineer', 'arivu@jdtech.com', '', date(2026, 1, 1)),
        ('EMP018', 'Sanjay', 'SCADA', 'SCADA & Python Developer', 'sanjay@jdtech.com', '', date(2026, 1, 1)),
        ('EMP019', 'Suguthan', 'PLC', 'PLC Engineer', 'suguthan@jdtech.com', '', date(2026, 1, 1)),
        ('EMP020', 'Shiva', 'PLC', 'PLC Engineer', 'shiva@jdtech.com', '', date(2026, 1, 1)),
        ('EMP021', 'Saravana Kumar', 'SCADA', 'SCADA Engineer', 'saravana@jdtech.com', '', date(2026, 1, 1)),
        ('EMP022', 'Naveen Kumar', 'SCADA', 'SCADA Engineer', 'naveen@jdtech.com', '', date(2026, 1, 1)),
        ('EMP023', 'Manobala', 'SCADA', 'SCADA Engineer', 'manobala@jdtech.com', '', date(2026, 1, 1)),
        ('EMP024', 'Vishnnu', 'SPC', 'SPC Engineer', 'vishnnu@jdtech.com', '', date(2026, 1, 1)),
        ('EMP025', 'Karthikeyan', 'Sales', 'Sales & SCADA Engineer', 'karthikeyan@jdtech.com', '', date(2026, 1, 1)),
        ('EMP026', 'Dinesh', 'Sales', 'Sales Executive', 'dinesh@jdtech.com', '', date(2026, 1, 1)),
        ('EMP027', 'Aravinth', 'SCADA,PLC', 'SCADA,PLC  Engineer', 'aravinth@jdtech.com', '', date(2026, 1, 1)),
    ]
    for eid, name, dept, pos, email, phone, jd in samples:
        if not Employee.query.filter_by(employee_id=eid).first():
            db.session.add(Employee(employee_id=eid, full_name=name, department=dept,
                                    position=pos, email=email, phone=phone, join_date=jd))
    
    # Seed default locations
    default_sites = ['Office', 'MRF TC1', 'ATG VIZAG', 'JK TYRES', 'MRF TCC', 'MRF TCR1 ZONE 2',
                     'MRF PDY PCR ZONE 1', 'MRF TCR1 CURING', 'MRF PONDY TBR', 'ARGC MVRLA']
    for site in default_sites:
        if not Location.query.filter_by(name=site).first():
            db.session.add(Location(name=site))
            
    db.session.commit()

# ── Auth decorators ────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Admin access required.', 'error')
            return redirect(url_for('my_attendance'))
        return f(*args, **kwargs)
    return decorated

# ── Haversine Distance (meters between two GPS points) ────────────────────────
import math

def haversine_distance(lat1, lon1, lat2, lon2):
    """Returns distance in meters between two GPS coordinates."""
    R = 6_371_000  # Earth radius in metres
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi       = math.radians(lat2 - lat1)
    dlambda    = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

def get_local_ip():
    import socket
    # Try a dummy socket connection to resolve the active primary network interface IP
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.settimeout(0)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        if ip and not ip.startswith('127.'):
            return ip
    except Exception:
        pass
    
    # Fallback to hostname lookup
    try:
        hostname = socket.gethostname()
        ip = socket.gethostbyname(hostname)
        if ip and not ip.startswith('127.'):
            return ip
    except Exception:
        pass
        
    # Second fallback: scan host addresses
    try:
        for ip in socket.gethostbyname_ex(socket.gethostname())[2]:
            if not ip.startswith('127.'):
                return ip
    except Exception:
        pass
        
    return '127.0.0.1'

# ── Secure Public Tunneling ───────────────────────────────────────────────────

import threading
import subprocess
import re
import time


public_url = None
_tunnel_process = None  # Module-level ref prevents garbage collection

def start_public_tunnel():
    global public_url, _tunnel_process
    time.sleep(1.5)

    SSH_OPTS = [
        "-C",                              # Enable SSH compression for faster asset loading
        "-o", "StrictHostKeyChecking=no",
        "-o", "ServerAliveInterval=10",    # Send keepalive ping every 10s (faster drop detection)
        "-o", "ServerAliveCountMax=2",     # Drop connection after 2 missed pings (20s total)
        "-o", "ExitOnForwardFailure=yes",  # Fail fast if port forward can't bind
    ]

    # Prioritize pinggy.io and serveo.net for speed and stability
    SERVICES = [
        ("free.pinggy.io",
         ["-p", "443", "-R", "0:127.0.0.1:5001"],
         r'https://[a-zA-Z0-9.-]+\.pinggy\.link'),
        ("serveo.net",
         ["-R", "80:127.0.0.1:5001"],
         r'https://[a-zA-Z0-9.-]+\.serveo\.net'),
        ("nokey@localhost.run",
         ["-R", "80:127.0.0.1:5001"],
         r'https://([a-zA-Z0-9-]+)\.(?:lhr\.life|lhrtunnel\.link|lhr\.rocks)'),
    ]

    def run_tunnel(host, ssh_args, url_pattern):
        global public_url, _tunnel_process
        try:
            cmd = ["ssh"] + SSH_OPTS + ssh_args + [host]
            process = subprocess.Popen(
                cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1)
            _tunnel_process = process

            for line in iter(process.stdout.readline, ''):
                match = re.search(url_pattern, line)
                if match:
                    public_url = match.group(0)
                    print("=" * 64)
                    print(f"  PUBLIC INTERNET LINK ACTIVE  [{host}]")
                    print(f"  URL : {public_url}")
                    print("=" * 64)
                    break

            # Block here — keeps thread alive while SSH session is open
            for _ in iter(process.stdout.readline, ''):
                pass

            process.wait()
        except Exception as e:
            print(f"  [Tunnel] {host} error: {e}")
        finally:
            _tunnel_process = None
            public_url = None  # Clear stale URL when session drops

    # ── Auto-reconnect loop — runs for the entire app lifetime ────────────────
    while True:
        for host, ssh_args, pattern in SERVICES:
            print(f"  [Tunnel] Connecting via {host}...")
            run_tunnel(host, ssh_args, pattern)
            print(f"  [Tunnel] {host} disconnected.")
        print("  [Tunnel] All services failed. Retrying in 10s...")
        time.sleep(10)




# ── Context Processor ──────────────────────────────────────────────────────────

@app.context_processor
def inject_network_info():
    global public_url
    local_ip = get_local_ip()
    active_url = public_url if public_url else f"http://{local_ip}:5001"
    return {
        'local_ip': local_ip,
        'public_url': public_url,
        'network_url': active_url
    }



@app.route('/')
def index():
    if 'user_id' in session:
        if session.get('role') == 'admin':
            return redirect(url_for('dashboard'))
        else:
            return redirect(url_for('my_attendance'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        pw   = hashlib.sha256(request.form['password'].encode()).hexdigest()
        user = User.query.filter_by(username=request.form['username'], password=pw).first()
        if user:
            session.update(user_id=user.id, username=user.username,
                           role=user.role, full_name=user.full_name,
                           employee_id=user.employee_id)
            if user.role == 'admin':
                return redirect(url_for('attendance_entry'))
            else:
                return redirect(url_for('my_attendance'))
        flash('Invalid username or password.', 'error')
    return render_template('login.html')

@app.route('/entry', methods=['GET', 'POST'])
@login_required
@admin_required
def attendance_entry():
    today       = date.today()
    employee_id = session.get('employee_id')
    existing    = None
    employee    = None

    if employee_id:
        employee = Employee.query.filter_by(employee_id=employee_id).first()
        existing = Attendance.query.filter_by(
            employee_id=employee_id, date=today).first()

    if request.method == 'POST' and employee_id:
        if not existing:
            now_time = datetime.now().strftime('%H:%M')
            db.session.add(Attendance(
                employee_id=employee_id,
                date=today,
                status='Present',
                check_in=now_time,
                marked_by=session['username']
            ))
            db.session.commit()
            flash(f'\u2705 Attendance marked! Your entry time is {now_time}.', 'success')
        return redirect(url_for('dashboard'))

    return render_template('attendance_entry.html',
        today=today.isoformat(), existing=existing,
        employee=employee, employee_id=employee_id)

@app.route('/my-attendance', methods=['GET', 'POST'])
@login_required
def my_attendance():
    employee_id = session.get('employee_id')
    employee    = None
    records     = []
    today       = date.today()
    today_record = None

    if employee_id:
        employee = Employee.query.filter_by(employee_id=employee_id).first()
        if employee:
            if request.method == 'POST':
                existing = Attendance.query.filter_by(
                    employee_id=employee_id, date=today).first()
                if existing:
                    flash('Attendance already recorded for today!', 'error')
                else:
                    status   = request.form.get('status', 'Present')
                    location = request.form.get('location', '')
                    notes    = request.form.get('notes', '')
                    now_time = datetime.now().strftime('%H:%M')

                    if status == 'Absent':
                        location = ''

                    # ── Geo-fence validation (employees only, skip WFH/Absent) ──
                    GEO_SKIP = {'WFH', 'Absent', ''}
                    if session.get('role') == 'employee' and location not in GEO_SKIP:
                        emp_lat  = request.form.get('emp_lat', '').strip()
                        emp_lng  = request.form.get('emp_lng', '').strip()
                        geo_skip = request.form.get('geo_skip', '').strip()  # 'no_coords' flag

                        if geo_skip != 'no_coords':
                            # Coordinates were NOT captured — browser denied GPS
                            if not emp_lat or not emp_lng:
                                flash('📍 Location access is required to mark attendance. Please allow GPS and try again.', 'error')
                                return redirect(url_for('my_attendance'))

                            try:
                                emp_lat = float(emp_lat)
                                emp_lng = float(emp_lng)
                            except ValueError:
                                flash('❌ Invalid GPS coordinates received. Please try again.', 'error')
                                return redirect(url_for('my_attendance'))

                            # Look up the site's saved coordinates
                            site = Location.query.filter_by(name=location).first()
                            if site and site.latitude is not None and site.longitude is not None:
                                distance = haversine_distance(emp_lat, emp_lng, site.latitude, site.longitude)
                                radius   = site.radius_meters or 200
                                if distance > radius:
                                    dist_display = int(distance)
                                    flash(
                                        f'🚫 Geo-fence check failed: You are {dist_display}m away from "{location}" '
                                        f'(allowed: {radius}m). Please reach the site and try again.',
                                        'error'
                                    )
                                    return redirect(url_for('my_attendance'))
                            # If site has no coordinates saved → allow (admin hasn't set them yet)

                    db.session.add(Attendance(
                        employee_id=employee_id,
                        date=today,
                        status=status,
                        check_in=now_time,
                        location=location,
                        notes=notes,
                        marked_by=session.get('username', 'Self')
                    ))
                    db.session.commit()
                    flash('✅ Attendance recorded successfully!', 'success')
                return redirect(url_for('my_attendance'))

            records = Attendance.query.filter_by(employee_id=employee_id)\
                .order_by(Attendance.date.desc()).limit(60).all()
            today_record = Attendance.query.filter_by(
                employee_id=employee_id, date=today).first()

    present_count = sum(1 for r in records if r.status == 'Present')
    absent_count  = sum(1 for r in records if r.status == 'Absent')
    halfday_count = sum(1 for r in records if r.status == 'Half Day')

    sites = [loc.name for loc in Location.query.order_by(Location.name).all()]

    return render_template('my_attendance.html',
        employee=employee, records=records, today=today.isoformat(),
        today_record=today_record, present_count=present_count,
        absent_count=absent_count, halfday_count=halfday_count,
        sites=sites)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
@admin_required
def dashboard():
    today         = date.today()
    total_emp     = Employee.query.filter_by(active=True).count()
    present_today = Attendance.query.filter_by(date=today, status='Present').count()
    absent_today  = Attendance.query.filter_by(date=today, status='Absent').count()
    not_marked    = total_emp - present_today - absent_today

    trend = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        p = Attendance.query.filter_by(date=d, status='Present').count()
        trend.append({'date': d.isoformat(), 'present': p})

    recent_q = (
        db.session.query(Attendance, Employee)
        .join(Employee, Attendance.employee_id == Employee.employee_id)
        .order_by(Attendance.date.desc(), Attendance.id.desc())
        .limit(10).all()
    )
    recent = [{
        'employee_id': a.employee_id, 'date': a.date.isoformat(), 'status': a.status,
        'location': a.location or '',
        'full_name': e.full_name, 'department': e.department
    } for a, e in recent_q]

    return render_template('dashboard.html',
        total=total_emp, present=present_today, absent=absent_today,
        not_marked=not_marked, trend=trend, recent=recent, today=today.isoformat())

@app.route('/api/dashboard/stats')
@login_required
@admin_required
def dashboard_stats():
    today         = date.today()
    total_emp     = Employee.query.filter_by(active=True).count()
    present_today = Attendance.query.filter_by(date=today, status='Present').count()
    absent_today  = Attendance.query.filter_by(date=today, status='Absent').count()
    not_marked    = total_emp - present_today - absent_today

    recent_q = (
        db.session.query(Attendance, Employee)
        .join(Employee, Attendance.employee_id == Employee.employee_id)
        .order_by(Attendance.date.desc(), Attendance.id.desc())
        .limit(10).all()
    )
    recent = [{
        'employee_id': a.employee_id, 'date': a.date.isoformat(), 'status': a.status,
        'location': a.location or '',
        'full_name': e.full_name, 'department': e.department
    } for a, e in recent_q]

    return jsonify({
        'total': total_emp,
        'present': present_today,
        'absent': absent_today,
        'not_marked': not_marked,
        'recent': recent
    })
    

@app.route('/mark', methods=['GET', 'POST'])
@login_required
@admin_required
def mark_attendance():
    today = date.today()
    if request.method == 'POST':
        sel_date  = date.fromisoformat(request.form.get('date', today.isoformat()))
        employees = Employee.query.filter_by(active=True).order_by(
            Employee.department, Employee.full_name).all()
        updated = 0
        for emp in employees:
            status    = request.form.get(f'status_{emp.employee_id}')
            check_in  = request.form.get(f'checkin_{emp.employee_id}', '')
            check_out = request.form.get(f'checkout_{emp.employee_id}', '')
            notes     = request.form.get(f'notes_{emp.employee_id}', '')
            location  = request.form.get(f'location_{emp.employee_id}', '')
            if status == 'Absent':
                location = ''
            if status:
                rec = Attendance.query.filter_by(
                    employee_id=emp.employee_id, date=sel_date).first()
                if rec:
                    rec.status, rec.check_in, rec.check_out = status, check_in, check_out
                    rec.notes, rec.marked_by, rec.location = notes, session['username'], location
                else:
                    db.session.add(Attendance(
                        employee_id=emp.employee_id, date=sel_date, status=status,
                        check_in=check_in, check_out=check_out,
                        notes=notes, marked_by=session['username'], location=location))
                updated += 1
        db.session.commit()
        flash(f'Attendance saved for {updated} employees on {sel_date}.', 'success')
        return redirect(url_for('mark_attendance'))

    selected_date = request.args.get('date', today.isoformat())
    sel_date_obj  = date.fromisoformat(selected_date)
    employees     = Employee.query.filter_by(active=True).order_by(
        Employee.department, Employee.full_name).all()
    existing      = {r.employee_id: r for r in
                     Attendance.query.filter_by(date=sel_date_obj).all()}
    sites = [loc.name for loc in Location.query.order_by(Location.name).all()]
    return render_template('mark_attendance.html',
        employees=employees, existing=existing, selected_date=selected_date,
        sites=sites)

@app.route('/employees')
@login_required
@admin_required
def employees():
    emps = Employee.query.filter_by(active=True).order_by(
        Employee.department, Employee.full_name).all()
    return render_template('employees.html', employees=emps)

@app.route('/employees/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_employee():
    if request.method == 'POST':
        try:
            jd = date.fromisoformat(request.form['join_date']) if request.form.get('join_date') else None
            db.session.add(Employee(
                employee_id=request.form['employee_id'],
                full_name=request.form['full_name'],
                department=request.form['department'],
                position=request.form.get('position', ''),
                email=request.form.get('email', ''),
                phone=request.form.get('phone', ''),
                join_date=jd))
            db.session.commit()
            flash('Employee added successfully!', 'success')
            return redirect(url_for('employees'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'error')
    return render_template('add_employee.html')

@app.route('/employees/delete/<emp_id>', methods=['POST'])
@login_required
@admin_required
def delete_employee(emp_id):
    emp = Employee.query.filter_by(employee_id=emp_id).first_or_404()
    emp.active = False
    db.session.commit()
    flash('Employee removed.', 'success')
    return redirect(url_for('employees'))

@app.route('/reports')
@login_required
@admin_required
def reports():
    filter_type = request.args.get('filter_type', 'monthly')
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    dept    = request.args.get('dept', 'All')

    if filter_type == 'custom' and start_date_str and end_date_str:
        try:
            start = date.fromisoformat(start_date_str)
            end = date.fromisoformat(end_date_str)
        except ValueError:
            start = date.today().replace(day=1)
            end = date.today()
            start_date_str = start.isoformat()
            end_date_str = end.isoformat()
    else:
        try:
            y, m    = map(int, month.split('-'))
            start   = date(y, m, 1)
            end     = date(y, m, calendar.monthrange(y, m)[1])
        except Exception:
            start   = date.today().replace(day=1)
            end     = date.today()
            month   = date.today().strftime('%Y-%m')
        start_date_str = start.isoformat()
        end_date_str = end.isoformat()

    end_day = (end - start).days + 1

    query = (
        db.session.query(
            Employee.employee_id, Employee.full_name, Employee.department,
            func.sum(case((Attendance.status == 'Present',  1), else_=0)).label('present_days'),
            func.sum(case((Attendance.status == 'Absent',   1), else_=0)).label('absent_days'),
            func.sum(case((Attendance.status == 'Half Day', 1), else_=0)).label('halfday_days'),
            func.count(Attendance.id).label('total_marked'),
        )
        .outerjoin(Attendance,
            (Employee.employee_id == Attendance.employee_id) &
            (Attendance.date.between(start, end)))
        .filter(Employee.active == True)
    )
    if dept != 'All':
        query = query.filter(Employee.department == dept)
    rows = query.group_by(
        Employee.employee_id, Employee.full_name, Employee.department
    ).order_by(Employee.department, Employee.full_name).all()

    depts = [r[0] for r in
             db.session.query(Employee.department).filter_by(active=True)
             .distinct().order_by(Employee.department).all()]

    
    loc_query = (
        db.session.query(Attendance.location, func.count(Attendance.id).label('count'))
        .join(Employee, Attendance.employee_id == Employee.employee_id)
        .filter(Employee.active == True)
        .filter(Attendance.date.between(start, end))
        .filter(Attendance.location != None)
        .filter(Attendance.location != '')
    )
    if dept != 'All':
        loc_query = loc_query.filter(Employee.department == dept)
    loc_rows = loc_query.group_by(Attendance.location).order_by(func.count(Attendance.id).desc()).all()

    return render_template('reports.html',
        rows=rows, month=month, start_date=start_date_str, end_date=end_date_str,
        filter_type=filter_type, dept=dept, depts=depts, end_day=end_day, loc_rows=loc_rows)

@app.route('/export')
@login_required
@admin_required
def export_csv():
    filter_type = request.args.get('filter_type', 'monthly')
    month = request.args.get('month', date.today().strftime('%Y-%m'))
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    dept    = request.args.get('dept', 'All')
    export_format = request.args.get('format', 'pdf')

    if filter_type == 'custom' and start_date_str and end_date_str:
        try:
            start = date.fromisoformat(start_date_str)
            end = date.fromisoformat(end_date_str)
        except ValueError:
            start = date.today().replace(day=1)
            end = date.today()
    else:
        try:
            y, m    = map(int, month.split('-'))
            start   = date(y, m, 1)
            end     = date(y, m, calendar.monthrange(y, m)[1])
        except Exception:
            start   = date.today().replace(day=1)
            end     = date.today()

    end_day = (end - start).days + 1

    query = (
        db.session.query(
            Employee.employee_id, Employee.full_name, Employee.department,
            Attendance.date, Attendance.status, Attendance.check_in,
            Attendance.location, Attendance.notes)
        .outerjoin(Attendance,
            (Employee.employee_id == Attendance.employee_id) &
            (Attendance.date.between(start, end)))
        .filter(Employee.active == True)
    )
    if dept != 'All':
        query = query.filter(Employee.department == dept)
    rows = query.order_by(Employee.department, Employee.full_name, Attendance.date).all()

    
    if export_format == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            ws.views.sheetView[0].showGridLines = True

            
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = "JD TECH — Attendance Report"
            title_cell.font = Font(name="Segoe UI", size=15, bold=True, color="00D4AA")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells('A2:H2')
            sub_cell = ws['A2']
            if filter_type == '__custom__' or filter_type == 'custom':
                sub_cell.value = f"Period: {start.isoformat()} to {end.isoformat()}  |  Department: {dept}  |  Generated: {date.today().isoformat()}"
            else:
                sub_cell.value = f"Month: {month}  |  Department: {dept}  |  Generated: {date.today().isoformat()}"
            sub_cell.font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
            sub_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 20
            ws.append([]) 

            
            headers = ['Emp ID', 'Full Name', 'Department', 'Date', 'Status', 'Check-In', 'Location', 'Notes']
            ws.append(headers)
            ws.row_dimensions[4].height = 24

            h_font = Font(name="Segoe UI", size=10, bold=True, color="00D4AA")
            h_fill = PatternFill(start_color="0D1F35", end_color="0D1F35", fill_type="solid")
            h_align = Alignment(horizontal="left", vertical="center")

            for col_idx in range(1, 9):
                cell = ws.cell(row=4, column=col_idx)
                cell.font = h_font
                cell.fill = h_fill
                cell.alignment = h_align

            
            border_side = Side(style='thin', color='1F2D45')
            thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

            f_reg = Font(name="Segoe UI", size=9, color="E2E8F0")
            f_present = Font(name="Segoe UI", size=9, bold=True, color="10B981")
            f_location = Font(name="Segoe UI", size=9, bold=True, color="00D4AA")

            fill_dark = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
            fill_darker = PatternFill(start_color="1A2235", end_color="1A2235", fill_type="solid")

            
            for idx, r in enumerate(rows):
                row_data = [
                    str(r.employee_id or ''),
                    str(r.full_name or ''),
                    str(r.department or ''),
                    str(r.date) if r.date else '',
                    str(r.status or ''),
                    str(r.check_in or '—'),
                    str(r.location or '—'),
                    str(r.notes or '')
                ]
                ws.append(row_data)
                curr_row = 5 + idx
                ws.row_dimensions[curr_row].height = 22
                curr_fill = fill_darker if idx % 2 == 1 else fill_dark

                for col_idx in range(1, 9):
                    cell = ws.cell(row=curr_row, column=col_idx)
                    cell.font = f_reg
                    cell.fill = curr_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

                    if col_idx == 5:
                        cell.font = f_present
                    elif col_idx == 7:
                        cell.font = f_location

            
            from openpyxl.utils import get_column_letter
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row in [1, 2]:
                        continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            filename_suffix = f"{start.isoformat()}_to_{end.isoformat()}" if filter_type == 'custom' else month
            return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=f'attendance_{filename_suffix}_{dept}.xlsx')
        except Exception as e:
            import traceback
            return f"<h3>Excel Export Error</h3><pre>{traceback.format_exc()}</pre>", 500

    
    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=landscape(A4),
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    
    title_style = ParagraphStyle('title', parent=styles['Normal'],
                                 fontSize=20, fontName='Helvetica-Bold',
                                 textColor=colors.HexColor('#00d4aa'),
                                 alignment=TA_CENTER, spaceAfter=4)
    sub_style   = ParagraphStyle('sub', parent=styles['Normal'],
                                 fontSize=11, fontName='Helvetica',
                                 textColor=colors.HexColor('#64748b'),
                                 alignment=TA_CENTER, spaceAfter=16)
    story.append(Paragraph('JD TECH \u2014 Attendance Report', title_style))
    if filter_type == 'custom':
        subtitle_text = f'Period: {start.isoformat()} to {end.isoformat()}  |  Department: {dept}  |  Generated: {date.today().isoformat()}'
    else:
        subtitle_text = f'Month: {month}  |  Department: {dept}  |  Generated: {date.today().isoformat()}'
    story.append(Paragraph(subtitle_text, sub_style))
    story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1f2d45'), spaceAfter=16))

    
    header = ['Emp ID', 'Full Name', 'Department', 'Date', 'Status', 'Check-In', 'Location', 'Notes']
    data   = [header]
    for r in rows:
        data.append([
            str(r.employee_id or ''),
            str(r.full_name or ''),
            str(r.department or ''),
            str(r.date) if r.date else '',
            str(r.status or ''),
            str(r.check_in or '—'),
            str(r.location or '—'),
            str(r.notes or ''),
        ])

    col_widths = [2.0*cm, 4.0*cm, 3.2*cm, 2.4*cm, 2.2*cm, 2.2*cm, 5.0*cm, 6.0*cm]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        
        ('BACKGROUND',   (0,0), (-1,0), colors.HexColor('#0d1f35')),
        ('TEXTCOLOR',    (0,0), (-1,0), colors.HexColor('#00d4aa')),
        ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,0), 9),
        ('TOPPADDING',   (0,0), (-1,0), 10),
        ('BOTTOMPADDING',(0,0), (-1,0), 10),
        
        ('BACKGROUND',   (0,1), (-1,-1), colors.HexColor('#111827')),
        ('TEXTCOLOR',    (0,1), (-1,-1), colors.HexColor('#e2e8f0')),
        ('FONTNAME',     (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',     (0,1), (-1,-1), 8),
        ('TOPPADDING',   (0,1), (-1,-1), 8),
        ('BOTTOMPADDING',(0,1), (-1,-1), 8),
        
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#111827'), colors.HexColor('#1a2235')]),
        
        ('GRID',         (0,0), (-1,-1), 0.5, colors.HexColor('#1f2d45')),
        ('ALIGN',        (0,0), (-1,-1), 'LEFT'),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        
        ('TEXTCOLOR',    (4,1), (4,-1), colors.HexColor('#10b981')),
        ('TEXTCOLOR',    (6,1), (6,-1), colors.HexColor('#00d4aa')),
    ]))
    story.append(tbl)

    
    story.append(Spacer(1, 16))
    story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1f2d45'), spaceAfter=8))
    story.append(Paragraph(f'Total records: {len(rows)}  |  JD TECH Attendance System  |  Confidential',
                           ParagraphStyle('footer', parent=styles['Normal'],
                                          fontSize=8, fontName='Helvetica',
                                          textColor=colors.HexColor('#64748b'),
                                          alignment=TA_CENTER)))

    doc.build(story)
    buf.seek(0)
    filename_suffix = f"{start.isoformat()}_to_{end.isoformat()}" if filter_type == 'custom' else month
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f'attendance_{filename_suffix}_{dept}.pdf')

@app.route('/users', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_users():
    if request.method == 'POST':
        username = request.form['username']
        if User.query.filter_by(username=username).first():
            flash(f"Username '{username}' already exists. Please choose a different one.", 'error')
        else:
            pw = hashlib.sha256(request.form['password'].encode()).hexdigest()
            emp_id = request.form.get('employee_id')
            if emp_id:
                emp_id = emp_id.strip()
            
            try:
                if emp_id:
                    existing_emp = Employee.query.filter_by(employee_id=emp_id).first()
                    if not existing_emp:
                        
                        new_emp = Employee(
                            employee_id=emp_id,
                            full_name=request.form['full_name'],
                            department=request.form.get('department') or 'General',
                            position='Employee',
                            email=request.form.get('email', ''),
                            join_date=date.today().isoformat()
                        )
                        db.session.add(new_emp)

                db.session.add(User(
                    username=username, password=pw,
                    role=request.form['role'], full_name=request.form['full_name'],
                    department=request.form.get('department', ''),
                    email=request.form.get('email', ''),
                    employee_id=emp_id if emp_id else None))
                
                db.session.commit()
                flash('User created successfully!', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'An error occurred: {str(e)}', 'error')
    users     = User.query.all()
    employees = Employee.query.filter_by(active=True).order_by(
        Employee.department, Employee.full_name).all()
    return render_template('users.html', users=users, employees=employees)

@app.route('/users/edit/<int:uid>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(uid):
    user = User.query.get_or_404(uid)
    if request.method == 'POST':
        username = request.form['username']
        
        existing_user = User.query.filter(User.username == username, User.id != uid).first()
        if existing_user:
            flash(f"Username '{username}' is already taken.", 'error')
        else:
            user.username = username
            user.role = request.form['role']
            user.full_name = request.form['full_name']
            user.department = request.form.get('department', '')
            user.email = request.form.get('email', '')
            
            
            new_password = request.form.get('password')
            if new_password:
                user.password = hashlib.sha256(new_password.encode()).hexdigest()
                
            
            emp_id = request.form.get('employee_id')
            if emp_id:
                emp_id = emp_id.strip()
            
            try:
                if emp_id:
                    existing_emp = Employee.query.filter_by(employee_id=emp_id).first()
                    if not existing_emp:
                        
                        new_emp = Employee(
                            employee_id=emp_id,
                            full_name=request.form['full_name'],
                            department=request.form.get('department') or 'General',
                            position='Employee',
                            email=request.form.get('email', ''),
                            join_date=date.today().isoformat()
                        )
                        db.session.add(new_emp)
                
                user.employee_id = emp_id if emp_id else None
                db.session.commit()
                flash('User updated successfully!', 'success')
                return redirect(url_for('manage_users'))
            except Exception as e:
                db.session.rollback()
                flash(f'An error occurred: {str(e)}', 'error')

    employees = Employee.query.filter_by(active=True).order_by(
        Employee.department, Employee.full_name).all()
    return render_template('edit_user.html', user=user, employees=employees)

@app.route('/users/delete/<int:uid>', methods=['POST'])
@login_required
@admin_required
def delete_user(uid):
    if uid == session['user_id']:
        flash('Cannot delete yourself.', 'error')
    else:
        User.query.filter_by(id=uid).delete()
        db.session.commit()
        flash('User deleted.', 'success')
    return redirect(url_for('manage_users'))

# ── Admin: Manage Location Geo-Coordinates ─────────────────────────────────────

@app.route('/admin/locations', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_locations():
    if request.method == 'POST':
        loc_id  = request.form.get('loc_id')
        lat     = request.form.get('latitude', '').strip()
        lng     = request.form.get('longitude', '').strip()
        radius  = request.form.get('radius_meters', '200').strip()
        loc = db.session.get(Location, loc_id)
        if loc:
            try:
                loc.latitude      = float(lat)  if lat  else None
                loc.longitude     = float(lng)  if lng  else None
                loc.radius_meters = int(radius) if radius else 200
                db.session.commit()
                flash(f'✅ Coordinates saved for "{loc.name}".', 'success')
            except ValueError:
                flash('❌ Invalid coordinates. Please enter valid numbers.', 'error')
        return redirect(url_for('manage_locations'))

    locations = Location.query.order_by(Location.name).all()
    return render_template('admin_locations.html', locations=locations)

# ── Public API: Location geo-data (for frontend GPS check) ─────────────────────

@app.route('/api/locations/geo')
@login_required
def api_locations_geo():
    """Returns geo-fence data for all locations so the frontend can pre-check."""
    locs = Location.query.all()
    data = {}
    for loc in locs:
        if loc.latitude is not None and loc.longitude is not None:
            data[loc.name] = {
                'lat':    loc.latitude,
                'lng':    loc.longitude,
                'radius': loc.radius_meters or 200
            }
    return jsonify(data)

@app.route('/api/locations/add', methods=['POST'])
@login_required
def add_location_api():

    data = request.get_json()
    if not data or 'name' not in data:
        return jsonify({'error': 'Invalid data'}), 400
    
    site_name = data['name'].strip()
    if not site_name:
        return jsonify({'error': 'Location name cannot be empty'}), 400
        
    existing = Location.query.filter_by(name=site_name).first()
    if existing:
        return jsonify({'success': True, 'name': existing.name, 'message': 'Location already exists'})
        
    try:
        new_loc = Location(name=site_name)
        db.session.add(new_loc)
        db.session.commit()
        return jsonify({'success': True, 'name': new_loc.name})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500



@app.route('/api-docs')
@login_required
def api_docs():
    return render_template('api_docs.html')



def get_current_user():
    
    token = request.headers.get('Authorization')
    if token:
        if token.startswith('Bearer '):
            token = token[7:]
        try:
            from itsdangerous import URLSafeSerializer
            serializer = URLSafeSerializer(app.secret_key)
            data = serializer.loads(token)
            user = User.query.get(data['user_id'])
            if user:
                return user
        except Exception:
            pass
            

    if 'user_id' in session:
        user = User.query.get(session['user_id'])
        if user:
            return user
            
    return None

def api_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        user = get_current_user()
        if not user:
            return jsonify({
                'error': 'Unauthorized access. Please login or provide a valid Bearer token in the Authorization header.'
            }), 401
        return f(user, *args, **kwargs)
    return decorated

def api_admin_required(f):
    @wraps(f)
    def decorated(user, *args, **kwargs):
        if user.role != 'admin':
            return jsonify({'error': 'Admin privileges required.'}), 403
        return f(user, *args, **kwargs)
    return decorated


@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.get_json() or {}
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'error': 'Username and password are required.'}), 400
        
    pw = hashlib.sha256(password.encode()).hexdigest()
    user = User.query.filter_by(username=username, password=pw).first()
    
    if not user:
        return jsonify({'error': 'Invalid username or password.'}), 401
        
    from itsdangerous import URLSafeSerializer
    serializer = URLSafeSerializer(app.secret_key)
    token = serializer.dumps({'user_id': user.id, 'username': user.username, 'role': user.role})
    
    return jsonify({
        'success': True,
        'token': token,
        'user': {
            'id': user.id,
            'username': user.username,
            'role': user.role,
            'full_name': user.full_name,
            'department': user.department or 'General',
            'email': user.email or '',
            'employee_id': user.employee_id
        }
    })  

@app.route('/api/employees', methods=['GET'])
@api_login_required
def api_get_employees(current_user):
    active_param = request.args.get('active')
    
    query = Employee.query
    if active_param is not None:
        is_active = active_param.lower() == 'true'
        query = query.filter(Employee.active == is_active)
        
    employees_list = query.order_by(Employee.department, Employee.full_name).all()
    
    result = [{
        'id': emp.id,
        'employee_id': emp.employee_id,
        'full_name': emp.full_name,
        'department': emp.department,
        'position': emp.position or '',
        'email': emp.email or '',
        'phone': emp.phone or '',
        'join_date': emp.join_date.isoformat() if emp.join_date else '',
        'active': emp.active
    } for emp in employees_list]
    
    return jsonify({'count': len(result), 'employees': result})

@app.route('/api/employees', methods=['POST'])
@api_login_required
@api_admin_required
def api_create_employee(current_user):
    data = request.get_json() or {}
    emp_id = data.get('employee_id')
    full_name = data.get('full_name')
    department = data.get('department')
    
    if not emp_id or not full_name or not department:
        return jsonify({'error': 'employee_id, full_name, and department are required fields.'}), 400
        
    emp_id = emp_id.strip()
    existing = Employee.query.filter_by(employee_id=emp_id).first()
    if existing:
        if existing.active:
            return jsonify({'error': f'Employee with ID {emp_id} already exists.'}), 400
        else:
            existing.active = True
            existing.full_name = full_name
            existing.department = department
            existing.position = data.get('position', existing.position)
            existing.email = data.get('email', existing.email)
            existing.phone = data.get('phone', existing.phone)
            join_date_str = data.get('join_date')
            if join_date_str:
                try:
                    existing.join_date = date.fromisoformat(join_date_str)
                except ValueError:
                    pass
            db.session.commit()
            return jsonify({
                'success': True,
                'message': 'Employee reactivated and updated.',
                'employee': {
                    'id': existing.id,
                    'employee_id': existing.employee_id,
                    'full_name': existing.full_name,
                    'department': existing.department,
                    'position': existing.position or '',
                    'email': existing.email or '',
                    'phone': existing.phone or '',
                    'join_date': existing.join_date.isoformat() if existing.join_date else '',
                    'active': existing.active
                }
            })
            
    join_date = None
    join_date_str = data.get('join_date')
    if join_date_str:
        try:
            join_date = date.fromisoformat(join_date_str)
        except ValueError:
            return jsonify({'error': 'Invalid join_date format. Use YYYY-MM-DD.'}), 400
    else:
        join_date = date.today()
        
    try:
        new_emp = Employee(
            employee_id=emp_id,
            full_name=full_name,
            department=department,
            position=data.get('position', ''),
            email=data.get('email', ''),
            phone=data.get('phone', ''),
            join_date=join_date,
            active=True
        )
        db.session.add(new_emp)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'employee': {
                'id': new_emp.id,
                'employee_id': new_emp.employee_id,
                'full_name': new_emp.full_name,
                'department': new_emp.department,
                'position': new_emp.position or '',
                'email': new_emp.email or '',
                'phone': new_emp.phone or '',
                'join_date': new_emp.join_date.isoformat() if new_emp.join_date else '',
                'active': new_emp.active
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to create employee: {str(e)}'}), 500

@app.route('/api/employees/<emp_id>', methods=['DELETE'])
@api_login_required
@api_admin_required
def api_delete_employee(current_user, emp_id):
    emp = Employee.query.filter_by(employee_id=emp_id).first()
    if not emp:
        return jsonify({'error': f'Employee with ID {emp_id} not found.'}), 404
        
    if not emp.active:
        return jsonify({'success': True, 'message': 'Employee is already inactive.'})
        
    try:
        emp.active = False
        db.session.commit()
        return jsonify({'success': True, 'message': f'Employee {emp_id} deactivated successfully.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to deactivate employee: {str(e)}'}), 500



@app.route('/api/attendance', methods=['GET'])
@api_login_required
def api_get_attendance(current_user):
    emp_id = request.args.get('employee_id')
    date_str = request.args.get('date')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    dept = request.args.get('department')
    
    
    if current_user.role != 'admin':
        if not current_user.employee_id:
            return jsonify({'error': 'Your user profile does not have an employee link. Contact administrator.'}), 400
        emp_id = current_user.employee_id
        
    query = db.session.query(Attendance, Employee).join(Employee, Attendance.employee_id == Employee.employee_id)
    
    if emp_id:
        query = query.filter(Attendance.employee_id == emp_id)
    if date_str:
        try:
            query = query.filter(Attendance.date == date.fromisoformat(date_str))
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400
    if start_date_str:
        try:
            query = query.filter(Attendance.date >= date.fromisoformat(start_date_str))
        except ValueError:
            return jsonify({'error': 'Invalid start_date format. Use YYYY-MM-DD.'}), 400
    if end_date_str:
        try:
            query = query.filter(Attendance.date <= date.fromisoformat(end_date_str))
        except ValueError:
            return jsonify({'error': 'Invalid end_date format. Use YYYY-MM-DD.'}), 400
    if dept:
        query = query.filter(Employee.department == dept)
        
    records = query.order_by(Attendance.date.desc(), Attendance.id.desc()).all()
    
    result = [{
        'id': a.id,
        'employee_id': a.employee_id,
        'full_name': e.full_name,
        'department': e.department,
        'date': a.date.isoformat(),
        'status': a.status,
        'check_in': a.check_in or '',
        'check_out': a.check_out or '',
        'location': a.location or '',
        'notes': a.notes or '',
        'marked_by': a.marked_by or ''
    } for a, e in records]
    
    return jsonify({'count': len(result), 'records': result})

@app.route('/api/attendance/mark', methods=['POST'])
@api_login_required
def api_mark_attendance(current_user):
    data = request.get_json() or {}
    emp_id = data.get('employee_id')
    
    if current_user.role != 'admin':
        if not current_user.employee_id:
            return jsonify({'error': 'Your user profile does not have an employee link. Contact administrator.'}), 400
        emp_id = current_user.employee_id
    elif not emp_id:
        if current_user.employee_id:
            emp_id = current_user.employee_id
        else:
            return jsonify({'error': 'Please specify employee_id.'}), 400
            
    employee = Employee.query.filter_by(employee_id=emp_id, active=True).first()
    if not employee:
        return jsonify({'error': f'Active employee with ID {emp_id} not found.'}), 404
        
    date_str = data.get('date')
    if date_str:
        try:
            attendance_date = date.fromisoformat(date_str)
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400
    else:
        attendance_date = date.today()
        
    status = data.get('status', 'Present')
    if status not in ['Present', 'Absent', 'Half Day']:
        return jsonify({'error': 'Status must be one of: Present, Absent, Half Day.'}), 400
        
    check_in = data.get('check_in')
    if not check_in:
        check_in = datetime.now().strftime('%H:%M')
        
    check_out = data.get('check_out')
    notes = data.get('notes', '')
    location = data.get('location', '')
    if status == 'Absent':
        location = ''
        
    existing = Attendance.query.filter_by(employee_id=emp_id, date=attendance_date).first()
    try:
        if existing:
            existing.status = status
            existing.check_in = check_in
            if check_out is not None:
                existing.check_out = check_out
            existing.location = location
            existing.notes = notes
            existing.marked_by = current_user.username
            db.session.commit()
            action = 'updated'
            rec = existing
        else:
            rec = Attendance(
                employee_id=emp_id,
                date=attendance_date,
                status=status,
                check_in=check_in,
                check_out=check_out,
                location=location,
                notes=notes,
                marked_by=current_user.username
            )
            db.session.add(rec)
            db.session.commit()
            action = 'created'
            
        return jsonify({
            'success': True,
            'action': action,
            'record': {
                'id': rec.id,
                'employee_id': rec.employee_id,
                'date': rec.date.isoformat(),
                'status': rec.status,
                'check_in': rec.check_in or '',
                'check_out': rec.check_out or '',
                'location': rec.location or '',
                'notes': rec.notes or '',
                'marked_by': rec.marked_by or ''
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to save attendance: {str(e)}'}), 500



@app.route('/api/locations', methods=['GET'])
@api_login_required
def api_get_locations(current_user):
    locs = Location.query.order_by(Location.name).all()
    return jsonify({
        'count': len(locs),
        'locations': [l.name for l in locs]
    })

@app.route('/api/locations', methods=['POST'])
@api_login_required
def api_create_location(current_user):
    data = request.get_json() or {}
    name = data.get('name')
    if not name:
        return jsonify({'error': 'Location name is required.'}), 400
        
    site_name = name.strip()
    if not site_name:
        return jsonify({'error': 'Location name cannot be empty.'}), 400
        
    existing = Location.query.filter_by(name=site_name).first()
    if existing:
        return jsonify({
            'success': True,
            'message': 'Location already exists.',
            'name': existing.name
        })
        
    try:
        new_loc = Location(name=site_name)
        db.session.add(new_loc)
        db.session.commit()
        return jsonify({
            'success': True,
            'name': new_loc.name
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/manifest.json')
def serve_manifest():
    return send_from_directory('static', 'manifest.json')

@app.route('/sw.js')
def serve_sw():
    response = make_response(send_from_directory('static', 'sw.js'))
    response.headers['Content-Type'] = 'application/javascript'
    response.headers['Service-Worker-Allowed'] = '/'
    return response

if __name__ == '__main__':
    with app.app_context():
        init_db()

    
    tunnel_thread = threading.Thread(target=start_public_tunnel, daemon=True)
    tunnel_thread.start()

    local_ip = get_local_ip()
    port = 5001

    print("=" * 64)
    print("  JD TECH Attendance System -- STARTING")
    print("=" * 64)
    print(f"  Local PC Link       : http://localhost:{port}")
    print(f"  MOBILE PHONE LINK   : http://{local_ip}:{port}")
    print("=" * 64)
    print("  * Connect your mobile phone to the same Wi-Fi network.")
    print("  * Type or scan this Mobile Link on your phone to open!")
    print("=" * 64)
    print()

    serve(app, host='0.0.0.0', port=port, threads=4)
